import openpyxl
import mysql.connector
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.chart import Series, LineChart, Reference

mydb = mysql.connector.connect(
    host="localhost",  # 数据库主机地址
    user="root",  # 数据库用户名
    passwd="20010518Hjy13",  # 数据库密码
    auth_plugin='mysql_native_password',
    database='peter'
)

myCursor = mydb.cursor()

print("输入文件名：")
fileName = input()

book = openpyxl.load_workbook(fileName + '.xlsx')
sheet = book.active

sheet.insert_cols(6, 2)
sheet.insert_cols(9, 2)
sheet.insert_cols(16, 10)

sheet.insert_rows(2, 1)

print("输入治疗日期：")
cureDateStr = input()

index = 0

cureDate = datetime.strptime(cureDateStr, '%Y/%m/%d')

for n in range(3, sheet.max_column):
    if sheet.cell(row=n, column=3).value <= cureDate <= sheet.cell(row=n + 1, column=3).value:
        index = n

sheet.cell(row=index, column=2).value = cureDate

for n in range(1, 12):
    sheet.cell(row=2, column=n + 3).value = n
for m in range(12, 25):
    sheet.cell(row=2, column=m + 4).value = m

for n in range(index - 1, 2, -1):
    sheet.cell(row=n, column=1).value = "C" + "-" + str(index - n)

for n in range(index, sheet.max_row + 1):
    sheet.cell(row=n, column=1).value = "C" + str(n - index + 1)

sheet.cell(row=1, column=6).value = "ΔALC比上次"
m = 0
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=5).value is not None:
        m = n - 1

    if sheet.cell(row=n, column=5).value is None:
        sheet.cell(row=n, column=6).value = ""
    elif sheet.cell(row=n - 1, column=5).value is None:
        sheet.cell(row=n, column=6).value = "=E" + str(n) + "-E" + str(m)
    else:
        sheet.cell(row=n, column=6).value = "=E" + str(n) + "-E" + str(n - 1)

sheet.cell(row=1, column=7).value = "ΔALC比C0"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=5).value is None:
        sheet.cell(row=n, column=7).value = ""
    else:
        sheet.cell(row=n, column=7).value = "=E" + str(n) + "-" + str(sheet.cell(row=index, column=5).value)

sheet.cell(row=1, column=9).value = "ΔANC比上次"
m = 0
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=8).value is not None:
        m = n - 1

    if sheet.cell(row=n, column=8).value is None:
        sheet.cell(row=n, column=9).value = ""
    elif sheet.cell(row=n - 1, column=8).value is None:
        sheet.cell(row=n, column=9).value = "=H" + str(n) + "-H" + str(m)
    else:
        sheet.cell(row=n, column=9).value = "=H" + str(n) + "-H" + str(n - 1)

sheet.cell(row=1, column=10).value = "ΔANC比C0"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=8).value is None:
        sheet.cell(row=n, column=10).value = ""
    else:
        sheet.cell(row=n, column=10).value = "=H" + str(n) + "-" + str(sheet.cell(row=index, column=8).value)

sheet.cell(row=1, column=16).value = "NLR"
for n in range(3, sheet.max_row + 1):
    if sheet.cell(row=n, column=5).value is None:
        sheet.cell(row=n, column=16).value = ""
    else:
        sheet.cell(row=n, column=16).value = "=H" + str(n) + "/E" + str(n)

sheet.cell(row=1, column=17).value = "ΔNLR比上次"
m = 0
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=16).value != "":
        m = n - 1

    if sheet.cell(row=n, column=16).value == "":
        sheet.cell(row=n, column=17).value = ""
    elif sheet.cell(row=n - 1, column=16).value == "":
        sheet.cell(row=n, column=17).value = "=P" + str(n) + "-P" + str(m)
    else:
        sheet.cell(row=n, column=17).value = "=P" + str(n) + "-P" + str(n - 1)

sheet.cell(row=1, column=18).value = "ΔNLR比基线"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=16).value == "":
        sheet.cell(row=n, column=18).value = ""
    else:
        sheet.cell(row=n, column=18).value = "=P" + str(n) + "-(H" + str(index) + "/E" + str(index) + ")"

sheet.cell(row=1, column=19).value = "dNLR"
for n in range(3, sheet.max_row + 1):
    if sheet.cell(row=n, column=4).value is None:
        sheet.cell(row=n, column=19).value = ""
    elif sheet.cell(row=n, column=8).value is None:
        sheet.cell(row=n, column=19).value = ""
    else:
        sheet.cell(row=n, column=19).value = "=H" + str(n) + "/(D" + str(n) + "-H" + str(n) + ")"

sheet.cell(row=1, column=20).value = "LMR"
for n in range(3, sheet.max_row + 1):
    if sheet.cell(row=n, column=12).value is None:
        sheet.cell(row=n, column=20).value = ""
    else:
        sheet.cell(row=n, column=20).value = "=E" + str(n) + "/L" + str(n)

m = 0
sheet.cell(row=1, column=21).value = "ΔLMR比上次"
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=20).value != "":
        m = n - 1

    if sheet.cell(row=n, column=20).value == "":
        sheet.cell(row=n, column=21).value = ""
    elif sheet.cell(row=n - 1, column=20).value == "":
        sheet.cell(row=n, column=21).value = "=T" + str(n) + "-T" + str(m)
    else:
        sheet.cell(row=n, column=21).value = "=T" + str(n) + "-T" + str(n - 1)

sheet.cell(row=1, column=22).value = "ΔLMR比基线"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=20).value == "":
        sheet.cell(row=n, column=22).value = ""
    else:
        sheet.cell(row=n, column=22).value = "=T" + str(n) + "-(E" + str(index) + "/L" + str(index) + ")"

sheet.cell(row=1, column=23).value = "PLR"
for n in range(3, sheet.max_row + 1):
    if sheet.cell(row=n, column=5).value is None:
        sheet.cell(row=n, column=23).value = ""
    else:
        sheet.cell(row=n, column=23).value = "=M" + str(n) + "/E" + str(n)

sheet.cell(row=1, column=24).value = "ΔPLR比上次"
m = 0
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=23).value != "":
        m = n - 1

    if sheet.cell(row=n, column=23).value == "":
        sheet.cell(row=n, column=24).value = ""
    elif sheet.cell(row=n - 1, column=23).value == "":
        sheet.cell(row=n, column=24).value = "=W" + str(n) + "-W" + str(m)
    else:
        sheet.cell(row=n, column=24).value = "=W" + str(n) + "-W" + str(n - 1)

sheet.cell(row=1, column=25).value = "ΔPLR比基线"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=23).value == "":
        sheet.cell(row=n, column=25).value = ""
    else:
        sheet.cell(row=n, column=25).value = "=W" + str(n) + "-(M" + str(index) + "/E" + str(index) + ")"

m = 0
sheet.cell(row=1, column=27).value = "ΔAMC比上次"
for n in range(4, sheet.max_row + 1):
    if sheet.cell(row=n - 1, column=12).value is not None:
        m = n - 1

    if sheet.cell(row=n, column=12).value is None:
        sheet.cell(row=n, column=27).value = ""
    elif sheet.cell(row=n - 1, column=12).value is None:
        sheet.cell(row=n, column=27).value = "=L" + str(n) + "-L" + str(m)
    else:
        sheet.cell(row=n, column=27).value = "=L" + str(n) + "-L" + str(n - 1)

sheet.cell(row=1, column=28).value = "ΔAMC比C0"
for n in range(index + 1, sheet.max_row + 1):
    if sheet.cell(row=n, column=12).value is None:
        sheet.cell(row=n, column=28).value = ""
    else:
        sheet.cell(row=n, column=28).value = "=L" + str(n) + "-" + str(sheet.cell(row=index, column=12).value)

sheet.column_dimensions[get_column_letter(3)].width = 13
sheet.column_dimensions[get_column_letter(15)].width = 13

# TODO: make charts
chart1 = LineChart()
data1 = Reference(sheet, min_col=4, min_row=3, max_row=sheet.max_row)
series1 = Series(data1, title='1')
chart1.append(series1)
data2 = Reference(sheet, min_col=5, min_row=3, max_row=sheet.max_row)
series2 = Series(data2, title='2')
chart1.append(series2)
data5 = Reference(sheet, min_col=8, min_row=3, max_row=sheet.max_row)
series3 = Series(data5, title='5')
chart1.append(series3)
data8 = Reference(sheet, min_col=11, min_row=3, max_row=sheet.max_row)
series4 = Series(data8, title='8')
chart1.append(series4)
data9 = Reference(sheet, min_col=12, min_row=3, max_row=sheet.max_row)
series5 = Series(data9, title='9')
chart1.append(series5)
data22 = Reference(sheet, min_col=26, min_row=3, max_row=sheet.max_row)
series6 = Series(data22, title='22')
chart1.append(series6)

chart1.title = '1'

xLabel = Reference(sheet, min_col=1, min_row=3, max_row=sheet.max_row)
chart1.set_categories(xLabel)

# Chart 2
chart2 = LineChart()
series1 = Series(data8, title='8')
chart2.append(series1)
series2 = Series(data9, title='9')
chart2.append(series2)
series3 = Series(data22, title='22')
chart2.append(series3)
chart2.title = '2'
chart2.set_categories(xLabel)

# Chart 3
chart3 = LineChart()
series1 = Series(data2, title='2')
chart3.append(series1)
series2 = Series(data5, title='5')
chart3.append(series2)
series3 = Series(data9, title='9')
chart3.append(series3)
series4 = Series(data22, title='22')
chart3.append(series4)
chart3.title = '3'
chart3.set_categories(xLabel)

# Chart 4
chart4 = LineChart()
series1 = Series(data5, title='5')
chart4.append(series1)
series2 = Series(data8, title='8')
chart4.append(series2)
series3 = Series(data22, title='22')
chart4.append(series3)
chart4.title = '4'
chart4.set_categories(xLabel)

# Chart 5
chart5 = LineChart()
data10 = Reference(sheet, min_col=13, min_row=3, max_row=sheet.max_row)
series1 = Series(data10, title='10')
chart5.append(series1)
data11 = Reference(sheet, min_col=14, min_row=3, max_row=sheet.max_row)
series2 = Series(data11, title='11')
chart5.append(series2)
chart5.title = '5'
chart5.set_categories(xLabel)

# Chart 6
chart6 = LineChart()
data3 = Reference(sheet, min_col=6, min_row=3, max_row=sheet.max_row)
series1 = Series(data3, title='3')
chart6.append(series1)
data4 = Reference(sheet, min_col=7, min_row=3, max_row=sheet.max_row)
series2 = Series(data4, title='4')
chart6.append(series2)
data6 = Reference(sheet, min_col=9, min_row=3, max_row=sheet.max_row)
series3 = Series(data6, title='6')
chart6.append(series3)
data7 = Reference(sheet, min_col=10, min_row=3, max_row=sheet.max_row)
series4 = Series(data7, title='7')
chart6.append(series4)
chart6.title = '6'
chart6.set_categories(xLabel)

# Chart 7
chart7 = LineChart()
data12 = Reference(sheet, min_col=16, min_row=3, max_row=sheet.max_row)
series1 = Series(data12, title='12')
chart7.append(series1)
data15 = Reference(sheet, min_col=19, min_row=3, max_row=sheet.max_row)
series2 = Series(data15, title='15')
chart7.append(series2)
data16 = Reference(sheet, min_col=20, min_row=3, max_row=sheet.max_row)
series3 = Series(data16, title='16')
chart7.append(series3)
chart7.title = '7'
chart7.set_categories(xLabel)

# Chart 8
chart8 = LineChart()
data19 = Reference(sheet, min_col=23, min_row=3, max_row=sheet.max_row)
series1 = Series(data19, title='19')
chart8.append(series1)
data20 = Reference(sheet, min_col=24, min_row=3, max_row=sheet.max_row)
series2 = Series(data20, title='20')
chart8.append(series2)
data21 = Reference(sheet, min_col=25, min_row=3, max_row=sheet.max_row)
series3 = Series(data21, title='21')
chart8.append(series3)
chart8.title = '8'
chart8.set_categories(xLabel)

# Chart 9
chart9 = LineChart()
series1 = Series(data3, title='3')
chart9.append(series1)
data23 = Reference(sheet, min_col=27, min_row=3, max_row=sheet.max_row)
series2 = Series(data23, title='23')
chart9.append(series2)
chart9.title = '9'
chart9.set_categories(xLabel)

position = ['A74', 'I74', 'P74', 'X74', 'A88', 'I88', 'P88', 'X88', 'A102']

chartList = [chart1, chart2, chart3, chart4, chart5, chart6, chart7, chart8, chart9]
for a, b in zip(iter(chartList), iter(position)):
    sheet.add_chart(a, b)

book.save(fileName + '结果' + '.xlsx')